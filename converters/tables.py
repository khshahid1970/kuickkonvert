"""PDF -> Excel.

LibreOffice has no native PDF-to-Calc export path (a PDF opens as a Draw or
Writer document internally, and neither can be saved as a spreadsheet), so
this tool takes a different, purpose-built approach: detect table
structures with pdfplumber and write each one to its own worksheet with
openpyxl, along with whatever plain text surrounds it on the page (a
heading above the table, summary totals below it, etc.) so nothing on the
page is silently dropped just because a table happens to sit on it. Pages
with no detected table still contribute their text, one line per row --
expect best results on PDFs with real ruled/structured tables rather than
free-form text or scanned images.
"""
import re

import pdfplumber
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .office import ConversionError

_INVALID_SHEET_CHARS = re.compile(r"[\[\]:\\/?*]")


def _safe_sheet_name(name: str) -> str:
    name = _INVALID_SHEET_CHARS.sub("-", name).strip() or "Sheet"
    return name[:31]


def _write_text_block(ws, text: str) -> bool:
    """Append each non-blank line of `text` as its own row. Returns True if
    anything was written."""
    lines = [ln for ln in (text or "").splitlines() if ln.strip()]
    for line in lines:
        ws.append([line])
    return bool(lines)


def convert_pdf_to_excel(input_path: str, out_dir: str) -> str:
    import os

    out_path = os.path.join(out_dir, "converted.xlsx")
    wb = Workbook()
    wb.remove(wb.active)  # replaced by real sheets below
    any_content = False

    try:
        with pdfplumber.open(input_path) as pdf:
            if len(pdf.pages) == 0:
                raise ConversionError("This PDF has no pages to convert.")

            for page_num, page in enumerate(pdf.pages, start=1):
                found_tables = page.find_tables()

                if not found_tables:
                    text = page.extract_text() or ""
                    lines = [ln for ln in text.splitlines() if ln.strip()]
                    if lines:
                        ws = wb.create_sheet(title=_safe_sheet_name(f"Page {page_num}"))
                        for line in lines:
                            ws.append([line])
                        ws.column_dimensions["A"].width = 100
                        any_content = True
                    continue

                # One sheet per page, containing -- in top-to-bottom reading
                # order -- any plain text above the first table, each table's
                # rows, any plain text between tables, and any trailing text
                # (e.g. statement summary totals) after the last table. A
                # blank row separates each block for readability. This
                # replaces the old behaviour of only ever capturing a page's
                # detected table(s) and silently dropping every other line of
                # text on that page -- confirmed 2026-08-27 on a real bank
                # statement PDF where the bank name, account holder's name,
                # account number and statement date (all plain text sitting
                # above the transaction table) were completely missing from
                # the converted spreadsheet, along with the transaction-total
                # summary lines sitting below the table.
                ws = wb.create_sheet(title=_safe_sheet_name(f"Page {page_num}"))
                max_cols = 1
                prev_bottom = 0.0
                sheet_has_content = False

                for table in sorted(found_tables, key=lambda t: t.bbox[1]):
                    top = table.bbox[1]
                    if top > prev_bottom + 1:
                        region = page.crop((0, prev_bottom, page.width, top))
                        if _write_text_block(ws, region.extract_text()):
                            ws.append([])
                            sheet_has_content = True

                    table_data = table.extract() or []
                    for row in table_data:
                        ws.append(["" if c is None else str(c) for c in row])
                    if table_data:
                        max_cols = max(max_cols, max(len(r) for r in table_data))
                        ws.append([])
                        sheet_has_content = True

                    prev_bottom = table.bbox[3]

                if prev_bottom < page.height - 1:
                    region = page.crop((0, prev_bottom, page.width, page.height))
                    if _write_text_block(ws, region.extract_text()):
                        sheet_has_content = True

                if sheet_has_content:
                    for col_idx in range(1, max_cols + 1):
                        ws.column_dimensions[get_column_letter(col_idx)].width = 22
                    any_content = True
                else:
                    wb.remove(ws)

    except ConversionError:
        raise
    except Exception as exc:
        raise ConversionError(
            "Could not read this PDF. It may be corrupted, password-protected, "
            "or a scanned image without extractable text."
        ) from exc

    if not any_content:
        raise ConversionError(
            "No text or tables could be found in this PDF -- it may be a scanned "
            "image. Try PDF OCR first (coming soon), or a text-based PDF."
        )

    wb.save(out_path)
    return out_path
